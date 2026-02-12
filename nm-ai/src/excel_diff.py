from copy import copy
import json
import logging
import os

import pandas as pd
from openpyxl.utils import get_column_letter

from src.db_ops import prep_queries_from_sql_files, run_query

logger = logging.getLogger(__name__)


TYPE_MAP = {
    "str": str,
    "int": int,
}


def _load_excel_config():
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    config_path = os.path.join(base_dir, "config", "excel_diff.json")
    with open(config_path, 'r', encoding='utf-8') as f:
        raw = json.load(f)
    return {
        "query_to_sheet_name": raw.get("query_to_sheet_name", {}),
        "sheet_field_translations": raw.get("sheet_field_translations", {}),
        "sheet_field_type_conversions": {
            sheet: {
                field: TYPE_MAP.get(type_name, str)
                for field, type_name in mapping.items()
            }
            for sheet, mapping in raw.get("sheet_field_type_conversions", {}).items()
        },
    }


_CONFIG = _load_excel_config()
QUERY_TO_SHEET_NAME = _CONFIG["query_to_sheet_name"]
SHEET_FIELD_TRANSLATIONS = _CONFIG["sheet_field_translations"]
SHEET_FIELD_TYPE_CONVERSIONS = _CONFIG["sheet_field_type_conversions"]


def compare_sheets(old_df, new_df, key_cols):
    merged = new_df.merge(old_df, on=key_cols, how='outer', indicator=True, suffixes=('_new', '_old'))
    added = merged[merged["_merge"] == "left_only"].drop(columns=["_merge"])
    removed = merged[merged["_merge"] == "right_only"].drop(columns=["_merge"])
    both = merged[merged["_merge"] == "both"]
    diffs = []
    for col in new_df.columns:
        if col in key_cols:
            continue
        mask = both[f"{col}_new"] != both[f"{col}_old"]
        if mask.any():
            tmp = both.loc[mask, key_cols + [f"{col}_old", f"{col}_new"]]
            diffs.append((col, tmp))
    return added, removed, diffs


def get_clean_header_from_worksheet(ws):
    header = []
    for cell in ws[1]:
        v = cell.value
        if isinstance(v, str):
            v = v.strip()
            if not v:
                continue
        if v is None:
            continue
        header.append(v)

    return header


def populate_deletions_sheet(wb, additional_results, verbose=False):
    """Populate the Deletions sheet with current deletions data."""
    if 'ai_current_deletions' not in additional_results:
        if verbose:
            logger.info("No 'ai_current_deletions' key found in additional_results")
            logger.info("Available keys: %s", list(additional_results.keys()))
        return

    deletions_df = additional_results['ai_current_deletions']

    if deletions_df.empty:
        if verbose:
            logger.info("Current deletions dataframe is empty")
        return

    if verbose:
        logger.info("Found %s deletion records", len(deletions_df))

    if "Deletions" not in wb.sheetnames:
        if verbose:
            logger.warning("Warning: 'Deletions' sheet not found in workbook")
            logger.info("Available sheets: %s", wb.sheetnames)
        return

    deletions_ws = wb["Deletions"]

    row1_values = [cell.value for cell in deletions_ws[1] if cell.value is not None]
    row2_values = [cell.value for cell in deletions_ws[2] if cell.value is not None]

    if len(row1_values) == 1 and len(row2_values) > 1:
        if verbose:
            logger.info("Detected title row in row 1, using row 2 as headers")
        header_row = 2
        data_start_row = 3
        header = []
        for cell in deletions_ws[2]:
            v = cell.value
            if isinstance(v, str):
                v = v.strip()
                if not v:
                    continue
            if v is None:
                continue
            header.append(v)
    else:
        if verbose:
            logger.info("Using row 1 as headers")
        header_row = 1
        data_start_row = 2
        header = get_clean_header_from_worksheet(deletions_ws)

    ai_code_col = None
    school_name_col = None
    reason_col = None

    for idx, col in enumerate(header, 1):
        if col == "AI Code":
            ai_code_col = idx
        elif col == "School Name":
            school_name_col = idx
        elif col == "Reason for Deletion":
            reason_col = idx

    if ai_code_col is None or school_name_col is None:
        if verbose:
            logger.warning("Required columns 'AI Code' or 'School Name' not found in Deletions sheet")
            logger.info("Available columns: %s", header)
        return

    if deletions_ws.max_row > data_start_row - 1:
        deletions_ws.delete_rows(data_start_row, deletions_ws.max_row - data_start_row + 1)

    def _first_value(row, columns):
        for col in columns:
            if col in row and pd.notna(row[col]):
                return str(row[col])
        return ''

    records_written = 0
    for row_idx, (_, row) in enumerate(deletions_df.iterrows(), start=data_start_row):
        ai_code = _first_value(row, ['ai_code', 'AI_CODE', 'ai_cd', 'AI_CD'])

        if ai_code:
            deletions_ws.cell(row=row_idx, column=ai_code_col, value=ai_code)

        school_name = _first_value(
            row,
            ['name', 'NAME', 'school_name', 'SCHOOL_NAME', 'ai_name', 'AI_NAME'],
        )

        if school_name:
            deletions_ws.cell(row=row_idx, column=school_name_col, value=school_name)

        if reason_col:
            deletions_ws.cell(row=row_idx, column=reason_col, value="")

        records_written += 1

        if verbose and row_idx <= data_start_row + 2:
            print(f"Row {row_idx}: AI Code='{ai_code}', School Name='{school_name}'")
            if row_idx == data_start_row:
                print(f"  Available columns in this row: {list(row.index)}")

    if verbose:
        print(f"Populated Deletions sheet with {records_written} records starting from row {data_start_row}")


def process_excel_diff(conn, args, target_month, old_sheets, wb, additional_results=None, email_statsD=None):
    """Process Excel diffing for merges and deletions queries."""
    diff_report = []
    detail_dfs = []
    if additional_results is None:
        additional_results = {}
    if email_statsD is None:
        email_statsD = {}

    merges_deletions_diff_query_allows = ['ai_domestic_trunc', 'ai_intl_trunc', 'ai_char_diffs']

    queries = prep_queries_from_sql_files("ai_diff_queries")
    queries_dict = {}

    for k, v in queries:
        if k in merges_deletions_diff_query_allows:
            if args.verbose:
                print(f"Running Excel diff query for {k}")
            queries_dict[k] = run_query(conn, v)

    if args.verbose:
        print("Excel diff queries executed.")
        print("QUERIES_DICT:")
        for k, v in queries_dict.items():
            print(f"  {k}: {v.shape}")

    if args.verbose:
        print("Comparing new data with old data...")

    for sql_name, new_df in queries_dict.items():
        actual_name = QUERY_TO_SHEET_NAME.get(sql_name, sql_name)
        if not actual_name or actual_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{actual_name}' not found in {args.infile}")

        old_df = old_sheets.get(actual_name, pd.DataFrame())

        field_map = SHEET_FIELD_TRANSLATIONS.get(actual_name, {})
        if field_map and args.verbose:
            print(f"Applying column-name mapping for '{actual_name}': {field_map}")
        if field_map:
            new_df = new_df.rename(columns=field_map)

        type_map = SHEET_FIELD_TYPE_CONVERSIONS.get(actual_name, {})
        for df in (old_df, new_df):
            for col, dtype in type_map.items():
                if col in df.columns:
                    df[col] = df[col].astype(dtype)

        ws = wb[actual_name]
        header = get_clean_header_from_worksheet(ws)

        if list(new_df.columns) != header:
            raise ValueError(
                f"Column mismatch in '{actual_name}':\n"
                f"  expected {header}\n"
                f"  got      {list(new_df.columns)}"
            )

        if actual_name == "Character replacements":
            key_cols = ["AI Code", "Attribute"]
        elif actual_name == "Truncations":
            key_cols = ["AI Code", "Truncated Attribute"]
        elif actual_name == "Truncations - International":
            key_cols = ["Ai Code", "Field"]
        else:
            key_cols = [header[0]]

        for df in (old_df, new_df):
            for k in key_cols:
                if k in df.columns:
                    df[k] = df[k].astype(str)

        if not old_df.empty:
            added, removed, diffs = compare_sheets(old_df, new_df, key_cols)

            changed_keys = set()
            for _, dfchg in diffs:
                for _, row in dfchg[key_cols].iterrows():
                    changed_keys.add(tuple(row[col] for col in key_cols))
            row_changes = len(changed_keys)

            if args.detail_report:
                if not added.empty:
                    added_clean = added.copy()
                    cols_to_keep = [
                        col for col in added_clean.columns if col.endswith('_new') or col in key_cols
                    ]
                    added_clean = added_clean[cols_to_keep]
                    added_clean.columns = [
                        col.replace('_new', '') if col.endswith('_new') else col for col in added_clean.columns
                    ]

                    if actual_name == "Truncations - International" and {
                        'Ai Code',
                        'Field',
                    }.issubset(added_clean.columns):
                        added_clean = added_clean.sort_values(['Ai Code', 'Field'])

                    detail_dfs.append((f"{actual_name}_ADDED", added_clean))
                    if args.verbose:
                        print(f"\n[{actual_name}] ADDED ({len(added_clean)} rows):")
                        print(added_clean.to_string(index=False))
                if not removed.empty:
                    removed_clean = removed.copy()
                    cols_to_keep = [
                        col for col in removed_clean.columns if col.endswith('_old') or col in key_cols
                    ]
                    removed_clean = removed_clean[cols_to_keep]
                    removed_clean.columns = [
                        col.replace('_old', '') if col.endswith('_old') else col for col in removed_clean.columns
                    ]

                    if actual_name == "Truncations - International" and {
                        'Ai Code',
                        'Field',
                    }.issubset(removed_clean.columns):
                        removed_clean = removed_clean.sort_values(['Ai Code', 'Field'])

                    detail_dfs.append((f"{actual_name}_REMOVED", removed_clean))
                    if args.verbose:
                        print(f"\n[{actual_name}] REMOVED ({len(removed_clean)} rows):")
                        print(removed_clean.to_string(index=False))
                for col, diff_df in diffs:
                    if actual_name == "Truncations - International" and {
                        'Ai Code',
                        'Field',
                    }.issubset(diff_df.columns):
                        diff_df = diff_df.sort_values(['Ai Code', 'Field'])

                    detail_dfs.append((f"{actual_name}_CHG_{col}", diff_df))
                    if args.verbose:
                        print(f"\n[{actual_name}] CHANGES in column '{col}' ({len(diff_df)} rows):")
                        print(diff_df.to_string(index=False))
            print(
                f"  + new rows: {len(added)}  - removed rows: {len(removed)}  * row-changes: {row_changes}"
            )
            diff_report.append({
                "sheet": actual_name,
                "added": len(added),
                "removed": len(removed),
                "row_changes": row_changes,
            })
            if sql_name == 'ai_domestic_trunc':
                email_statsD['ai_domestic_net'] = len(added) - len(removed)
                email_statsD['ai_domestic_added'] = len(added)
                email_statsD['ai_domestic_removed'] = len(removed)
                email_statsD['ai_domestic_row_changes'] = row_changes
                email_statsD['prev_total_domestic_diffs'] = len(old_df)
                email_statsD['total_ai_domestic_trunc'] = len(new_df)

            elif sql_name == 'ai_char_diffs':
                email_statsD['ai_char_diffs_net'] = len(added) - len(removed)
                email_statsD['ai_char_diffs_added'] = len(added)
                email_statsD['ai_char_diffs_removed'] = len(removed)
                email_statsD['ai_char_diffs_row_changes'] = row_changes
                email_statsD['prev_total_char_diffs'] = len(old_df)
                email_statsD['total_ai_char_diffs'] = len(new_df)

            elif sql_name == 'ai_intl_trunc':
                email_statsD['ai_intl_net'] = len(added) - len(removed)
                email_statsD['ai_intl_added'] = len(added)
                email_statsD['ai_intl_removed'] = len(removed)
                email_statsD['ai_intl_row_changes'] = row_changes
                email_statsD['prev_total_intl_diffs'] = len(old_df)
                email_statsD['total_ai_intl_trunc'] = len(new_df)
            else:
                print("WARNING: Unrecognized SQL name")
        else:
            print(
                f"  No existing data to compare in '{actual_name}', treating all {len(new_df)} as new."
            )
            diff_report.append({
                "sheet": actual_name,
                "added": len(new_df),
                "removed": 0,
                "row_changes": 0,
            })

        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        for r, row in enumerate(new_df.itertuples(index=False), start=2):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)

    if additional_results:
        populate_deletions_sheet(wb, additional_results, args.verbose)

    merges_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == "merges":
            merges_sheet = wb[sheet_name]
            break

    if merges_sheet is not None and merges_sheet.max_row > 2:
        merges_sheet.delete_rows(3, merges_sheet.max_row - 2)

    outname = f"{target_month} NMSC AI Merges and Deletions.xlsx"
    outpath = os.path.join(args.outdir, outname)
    wb.save(outpath)
    logger.info("New workbook saved to %s", outpath)

    if additional_results:
        j_to_s_df = additional_results.get("ai_j_to_s_detail", pd.DataFrame())
        s_to_j_df = additional_results.get("ai_s_to_j_detail", pd.DataFrame())
        j_to_s_out = pd.DataFrame({
            "AI code": j_to_s_df.get("ai_code", pd.Series(dtype=object)).astype(str),
            "name": j_to_s_df.get("name", pd.Series(dtype=object)).astype(str),
        })
        s_to_j_out = pd.DataFrame({
            "AI code": s_to_j_df.get("ai_code", pd.Series(dtype=object)).astype(str),
            "name": s_to_j_df.get("name", pd.Series(dtype=object)).astype(str),
        })

        transition_path = os.path.join(args.outdir, f"{target_month} NMSC AI School Level Changes.xlsx")
        with pd.ExcelWriter(transition_path, engine="openpyxl") as writer:
            j_to_s_out.to_excel(writer, sheet_name="J to S", index=False)
            s_to_j_out.to_excel(writer, sheet_name="S to J", index=False)
        logger.info("J to S / S to J workbook saved to %s", transition_path)

    if args.detail_report and detail_dfs:
        detail_path = os.path.join(args.outdir, f"diff_details_{target_month}.xlsx")
        with pd.ExcelWriter(detail_path, engine="openpyxl") as det_writer:
            for sheet_name, df in detail_dfs:
                safe_name = sheet_name[:31]
                df.to_excel(det_writer, sheet_name=safe_name, index=False)

                ws = det_writer.book[safe_name]
                ws.freeze_panes = "A2"

                for idx, col in enumerate(df.columns, 1):
                    max_len = max(
                        df[col].astype(str).map(len).max(),
                        len(col),
                    ) + 2
                    ws.column_dimensions[get_column_letter(idx)].width = max_len
        logger.info("Detailed diff workbook saved to %s", detail_path)

    return diff_report, detail_dfs
